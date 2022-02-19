import os
import openpyxl as xl
from functions import update_output
# from openpyxl.utils import get_column_letter
import openpyxl.styles as styles


def operate_file(window, values, titles):
    folder_path = values['folder']
    if os.path.exists(folder_path):
        files = os.listdir(folder_path)
    else:
        return -1
    key = values['key'] if values['include'] else None
    print(files)

    for file in files:
        update_output(window, values, f'{file}开始处理')
        print(file)
        file_path = f'{folder_path}/{file}'
        inf = open_file(file_path, titles)
        write_file('汇总的.xlsx', inf, key, titles)
        update_output(window, values, f'{file}已处理完成并写入')


def open_file(file_path, titles: list = None):

    sheet = xl.load_workbook(file_path).active
    max_row = sheet.max_row
    max_col = sheet.max_column

    min_row = 2 if titles else 1

    inf = []
    single_inf = []
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row,
                               max_col=max_col, values_only=True):
        for cell in row:
            single_inf.append(cell)

        inf.append(single_inf)
        single_inf = []

    return inf


def write_file(file_path, inf,
               key: str = None,
               titles: list = None):
    print(key)
    if inf == -1:
        return None
    elif os.path.exists(file_path):
        workbook = xl.load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = xl.Workbook()
        sheet = workbook.active
        if titles:
            sheet.append(titles)

    if key and key in titles:
        key_index = titles.index(key) + 1
        keyin_col = key_index + 1
        _key_col = sheet.iter_cols(min_col=keyin_col, max_col=keyin_col,
                                   values_only=True)
        key_col = []
        for col in _key_col:
            for cell in col:
                key_col.append(cell)

        print(key_col)
    else:
        key_col = None
        key_index = 0
        keyin_col = None

    warning_style = styles.PatternFill(start_color='FF0000', fill_type='solid')
    for single_inf in inf:
        now_row = (sheet.max_row + 1)
        # max_col = get_column_letter(sheet.max_column)
        if key:
            sheet.append(single_inf)
            if single_inf[key_index] in key_col:
                sheet.cell(row=now_row, column=keyin_col).fill = warning_style
                print('123123123131231')
        else:
            sheet.append(single_inf)

    workbook.save('汇总的.xlsx')


__all__ = ['operate_file']
